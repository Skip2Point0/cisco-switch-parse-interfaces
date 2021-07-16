import re
import os
from openpyxl import Workbook
configfiles = []
wb = Workbook()


def search_configfiles():
    for name in os.listdir("."):
        if name.endswith(".txt"):
            # print(name)
            reg = re.compile(r'(\d+.\d+.\d+.\d+)_switchconfig.txt')
            files = reg.search(name)
            if files:
                configfiles.append(name)
    print(configfiles)


def calculate_interface(cnf):
    # global interface_dictionary
    for num_lines in range(len(cnf)):
        line = cnf[num_lines]
        reg = re.compile(r'interface\s(TenGigabitEthernet|GigabitEthernet|FastEthernet|Ethernet)(\d.\d.\d+)')
        interface = reg.search(line)
        reg = re.compile(r'interface\sPort-channel(\d+)')
        port_channel = reg.search(line)
        if interface:
            # print(interface)
            interfaces.append(interface.group(1) + interface.group(2))
            interface_line.append(num_lines)
            # interface_dictionary[interface.group(1) + interface.group(2)] = {}
        elif port_channel:
            po = "Po" + port_channel.group(1)
            interfaces.append(po)
            interface_line.append(num_lines)
            # interface_dictionary[po] = {}
        else:
            continue


def calculate_interface_config(cnf, int_line):
    # global interface_dictionary
    for index in range(len(int_line)):
        interface_structure = {"number": [],
                               "name": [],
                               "enabled": [],
                               "type": [],
                               "vlan": [],
                               "voiceVlan": [],
                               "allowedVlans": [],
                               "rstpEnabled": [],
                               "stpGuard": [],
                               "portChannel": [],
                               "portChannelMode": [],
                               "linkNegotiation": []}
        object_line = int_line[index]
        for line in cnf[object_line + 1:]:
            # print(line)
            if line != "!":
                # print(line)
                reg = re.compile(r'\sswitchport\saccess\svlan\s(\d+)')
                v = reg.search(line)
                reg = re.compile(r'\sswitchport\svoice\svlan\s(\d+)')
                vv = reg.search(line)
                reg = re.compile(r'\sswitchport\strunk\sallowed\svlan\s(\d+.\d+.\d+)')
                av = reg.search(line)
                reg = re.compile(r'\sswitchport\smode\s(access|trunk)')
                pm = reg.search(line)
                reg = re.compile(r'\sdescription\s(.*)')
                descr = reg.search(line)
                reg = re.compile(r'\s(shut)')
                shut = reg.search(line)
                reg = re.compile(r'\sspanning-tree\s(portfast)')
                portf = reg.search(line)
                reg = re.compile(r'\sspanning-tree\sbpduguard\s(\w+)')
                bpdug = reg.search(line)
                reg = re.compile(r'\schannel-group\s(\d+)\smode\s(\w+)')
                chgrp = reg.search(line)
                reg = re.compile(r'\snegotiation\s(\w+)')
                portsp = reg.search(line)
                interface_structure["number"] = interfaces[index]
                if descr:
                    # print(descr.group(1))
                    interface_structure["name"] = descr.group(1)
                elif pm:
                    # print(pm.group(1))
                    interface_structure["type"] = pm.group(1)
                elif v:
                    # print(av.group(1))
                    interface_structure["vlan"] = v.group(1)
                elif vv:
                    # print(av.group(1))
                    interface_structure["voiceVlan"] = vv.group(1)
                elif av:
                    interface_structure["allowedVlans"] = av.group(1)
                elif shut:
                    interface_structure["enabled"] = "shut"
                elif portf:
                    interface_structure["rstpEnabled"] = "true"
                elif bpdug:
                    interface_structure["stpGuard"] = "true"
                elif chgrp:
                    interface_structure["portChannel"] = chgrp.group(1)
                    interface_structure["portChannelMode"] = chgrp.group(2)
                elif portsp:
                    interface_structure["linkNegotiation"] = portf.group(1)

            else:
                # interface_structure["type"] = " "
                # interface_structure["vlan"] = " "
                # interface_structure["enabled"] = " "
                # interface_structure["rstpEnabled"] = " "
                interface_structure["number"] = interfaces[index]
                interface_dictionary.append(interface_structure)
                break
        # interface_dictionary.append(interface_structure)


def replace_novalue_with_whitespace(dictionary):
    # print(dictionary)
    for d in dictionary:
        # print(d)
        for a in d:
            # print(a)
            if d[a]:
                continue
            else:
                d[a] = ' '
    # print(dictionary)
    return dictionary


def generate_spread(ip, dictionary, counter):
    print(dictionary)
    ws1 = wb.create_sheet(ip, counter)
    # print(wb.sheetnames)
    ws1['A1'] = 'Name'
    ws1['B1'] = 'Description'
    ws1['C1'] = 'Status'
    ws1['D1'] = 'Type'
    ws1['E1'] = 'Vlan'
    ws1['F1'] = 'Voice Vlan'
    ws1['G1'] = 'Allowed Vlans'
    ws1['H1'] = 'Portfast'
    ws1['I1'] = 'BPDU Guard'
    ws1['J1'] = 'Port-Channel #'
    ws1['K1'] = 'Port-Channel Mode'
    ws1['L1'] = 'Link Speed Negotiation'
    lc = 2
    for d in dictionary:
        n_ind = 'A' + str(lc)
        d_ind = 'B' + str(lc)
        s_ind = 'C' + str(lc)
        tp_ind = 'D' + str(lc)
        v_ind = 'E' + str(lc)
        vv_ind = 'F' + str(lc)
        av_ind = 'G' + str(lc)
        p_ind = 'H' + str(lc)
        b_ind = 'I' + str(lc)
        pc_ind = 'J' + str(lc)
        pcm_ind = 'K' + str(lc)
        l_ind = 'L' + str(lc)

        ws1[n_ind] = d['number']
        ws1[d_ind] = d['name']
        ws1[s_ind] = d['enabled']
        ws1[tp_ind] = d['type']
        ws1[v_ind] = d['vlan']
        ws1[vv_ind] = d['voiceVlan']
        ws1[av_ind] = d['allowedVlans']
        ws1[p_ind] = d['rstpEnabled']
        ws1[b_ind] = d['stpGuard']
        ws1[pc_ind] = d['portChannel']
        ws1[pcm_ind] = d['portChannelMode']
        ws1[l_ind] = d['linkNegotiation']

        lc = lc + 1


search_configfiles()
ctr = 0
for f in configfiles:
    interface_line = []
    interfaces = []
    interface_dictionary = []
    with open(f) as g:
        config = g.read().splitlines()
    reg = re.compile(r'(\d+.\d+.\d+.\d+)_switchconfig.txt')
    ipandtext = reg.search(f)
    calculate_interface(config)
    calculate_interface_config(config, interface_line)
    # print(interface_line)
    # print(interfaces)
    generate_spread(ipandtext.group(1), replace_novalue_with_whitespace(interface_dictionary), ctr)
    # print(interface_dictionary)
    ctr = ctr + 1

wb.save('Switches and Ports.xlsx')
